import pandas as pd                           #Importing Required Libraries
import re
import openpyxl


def WrtFile(row , index , sheet):                                                      #Function for writing selected students to xlsx file
    (sheet.cell(row,1)).value = row-1
    if(row < 11):
        (sheet.cell(row,2)).value  = 'EE-0'+str(row-1)
    else:
        (sheet.cell(row,2)).value  = 'EE-'+str(row-1)
        
    (sheet.cell(row,3)).value  = df_in['UserId'][index]
    (sheet.cell(row,4)).value  = df_in['159_e_full_name'][index]
    (sheet.cell(row,5)).value  = df_in['159_e_father_or_guardian_or_spouse_name'][index]
    (sheet.cell(row,6)).value  = df_in['159_h_date_of_birth'][index]
    (sheet.cell(row,7)).value  = df_in['159_y_category'][index]
    (sheet.cell(row,8)).value  = df_in['159_d_physically_handicapped'][index]
    (sheet.cell(row,9)).value  = df_in['159_r_phone_number'][index]
    (sheet.cell(row,10)).value = df_in['159_d_email_id'][index]
    (sheet.cell(row,11)).value = df_in['159_y_seeking_phd_admission_under_category'][index]
    (sheet.cell(row,12)).value = ''
 
col = 1
row = 2
row_n = 2
rej_count = 0
wb = openpyxl.Workbook()  
wb_n =  openpyxl.Workbook()                                                      #Creatind Workbook Object
sheet = wb.active
sheet_n = wb_n.active
Headers = ['SL No.','Appl. Sl. No.','Appl. No.','Candidates Name',"Father's Name",'DOB','Category' ,'PD','Mobile','E-mail'	,'Full Time/ Part Time/ Sponsored','REMARKS']
for i in Headers:                                                              #Naming Of all Columns
    (sheet.cell(1,col)).value = i
    (sheet_n.cell(1,col)).value = i
    col = col+1    
    

m = re.compile('IIT', re.I)
n  =re.compile('Indian Institute Of Technology.*',re.I)
o = re.compile('i\.i\.t',re.I)

mtech = re.compile('m.*tech',re.I|re.S)
mtech1 = re.compile('master.* of technology',re.I)
btech =  re.compile('b.*tech',re.I|re.S)
btech1 = re.compile('Bachelor Of Technology' ,re.I)
be = re.compile('b\.?e' ,re.I)
be1 = re.compile('bachelor of engineering',re.I)
me = re.compile('m.?e',re.I)
me1 = re.compile('master.? of engineering',re.I)
msc = re.compile('m.?sc',re.I)
msc1 = re.compile('master of science',re.I)
mca = re.compile('mca',re.I)
mca1 = re.compile('Master of Computer Application.?',re.I)
mba = re.compile('mba',re.I)
mba1 = re.compile('Master of Business Administration' ,re.I)
mphil = re.compile('M.?Phil',re.I|re.S)
mphil1 = re.compile('Master of Philosophy',re.I)
ugc = re.compile('ugc.*net',re.I|re.S)
dbt = re.compile('dbt.*jrf',re.I|re.S)
g = re.compile('Gate',re.I)



#input_file = 'raw_info_from_form_input_file.csv'
input_file = 'EE_Phd_397.csv'                              
df_in = pd.read_csv(input_file,delimiter = ',')                               #Transferring data from csv file to pandas Datarame 
Invalid_degree = []
rejected_df = pd.DataFrame(columns = df_in.columns)
for ind in df_in.index:
    # print('New')
    # print()
    # print()
    check_flag = 0
    check_flag2 = 0
    check_flag3 = 0
    check_flag4 = 0
    check_flag5 = 0
                                                            
    age_exemp = 0
    gate_exemp = 0
    ugcpi,ugper,qcpi,qper,hscpi,hsper = [0]*6
    ugcpi_l,ugper_l,imper_l,hscpi_l,hsper_l = [0]*5                           #Set this for additional conditions on X,XII or UGCPI. You can set the passing value here. However it will be applied to all degrees
    
    Exam_flag = 0
    gate_qualify = 0
    gate_valid = 0
    
    Candidate = df_in['159_e_full_name'][ind]
    qfd = df_in['101_e_qualification_degree'][ind]
    IIT_check = df_in['101_y_name_and_place_of_institution_or_university'][ind]
    Exam = df_in['109_e_exam_name'][ind]
    Exam2 = df_in['105_e_exam_name'][ind]

    if(df_in['101_e_overall_percentage_of_marks_or_final_grade_point_average'][ind]<=10):
        qcpi = df_in['101_e_overall_percentage_of_marks_or_final_grade_point_average'][ind]
    elif(df_in['101_e_overall_percentage_of_marks_or_final_grade_point_average'][ind]>10):
        qper = df_in['101_e_overall_percentage_of_marks_or_final_grade_point_average'][ind]
    if(df_in['104_e_percentage_of_marks_or_final_grade_point_average'][ind]<=10):
        ugcpi = df_in['104_e_percentage_of_marks_or_final_grade_point_average'][ind]
    elif(df_in['104_e_percentage_of_marks_or_final_grade_point_average'][ind]>10):
        ugper = df_in['104_e_percentage_of_marks_or_final_grade_point_average'][ind]     
    imper =  df_in['103_e_percentage_of_marks_or_final_grade_point_average'][ind] 
    if(df_in['102_e_percentage_of_marks_or_final_grade_point_average'][ind]<=10):
        hscpi = df_in['102_e_percentage_of_marks_or_final_grade_point_average'][ind]
    elif(df_in['102_e_percentage_of_marks_or_final_grade_point_average'][ind]>10):
        hsper = df_in['102_e_percentage_of_marks_or_final_grade_point_average'][ind]       
    category = df_in['159_y_category'][ind] 
    gender =   df_in['159_r_gender'][ind] 
    age =   2020 - int(df_in['159_h_date_of_birth'][ind].strip()[-4:])
    Phd_category = df_in['159_y_seeking_phd_admission_under_category'][ind]
    
    if(category == 'SC' or category == 'ST'):
        qcpi_btech = 7.5
    else:
        qcpi_btech = 8
    
    
    
    if(bool(g.match(Exam))):
        if(pd.isna(df_in['109_o_valid_upto'][ind]) == False and df_in['109_o_valid_upto'][ind]!='' and df_in['109_o_valid_upto'][ind] != '--'):
            gate_qualify = 1
            if(int(df_in['109_o_valid_upto'][ind])>=2020):
                gate_valid = 1        
    if(not(pd.isna(Exam2))):
        if(bool(g.match(Exam2))):
            if(pd.isna(df_in['105_o_valid_upto'][ind]) == False and df_in['105_o_valid_upto'][ind]!='' and df_in['105_o_valid_upto'][ind] != '--'):
                gate_qualify = 1
                if(int(df_in['105_o_valid_upto'][ind])>=2020):
                    gate_valid = 1    
            elif(bool(ugc.search(Exam2)) or Exam2 == 'NBHM' or bool(dbt.search(Exam2))):
                Exam_flag = 1
        
    
    if(bool(ugc.search(Exam)) or Exam == 'NBHM' or bool(dbt.search(Exam))):
        Exam_flag = 1
        
        
    PD = df_in['159_d_physically_handicapped'][ind]

    
    
    if(m.match(IIT_check)!= None or n.search(IIT_check)!=None or o.match(IIT_check)!= None):
       
        IIT_flag = 1          #Checking if IIT
    else:
        IIT_flag = 0
        
    if(Phd_category!='Regular and Full Time' ):       #Removing Age criteria if experience>=2years
        age_exemp = 1
    
    if(Phd_category in ['Employed and Part Time' , 'Self -Financed' , 'Sponsored']):
        gate_exemp = 1
    elif(Phd_category == 'Project Staff'):
        gate_exemp = gate_qualify
    
    Mtech_check = bool(mtech.search(qfd))|bool(mtech1.search(qfd))            #Various checks for identifying degree
    MS_check    = (qfd.lower() == 'ms') or(qfd.lower() == 'm.s') or(qfd.lower() == 'm.s.') or (qfd.lower() == 'master of science')
    ME_check    = bool(me.match(qfd))|bool(me1.search(qfd))
    Btech_check = bool(btech.search(qfd))|bool(btech1.search(qfd))|bool(be.match(qfd))|bool(be1.search(qfd))
    #MCA_check   = bool(mca.match(qfd))|bool(mca1.search(qfd))
    MSC_check   = bool(msc.search(qfd))|bool(msc1.search(qfd))
    
    if(Mtech_check == False and MS_check == False and ME_check==False and Btech_check == False  and MSC_check == False):
        Invalid_degree.append([df_in['UserId'][ind],Candidate,qfd])
        rejected_df.loc[rej_count] = df_in.loc[ind]
        rej_count+=1
        
    if(Btech_check and IIT_flag):
        qcpi_btech = 7
        
        
    
    
    
    if(category == 'SC' or category == 'ST'):
        if((Mtech_check or MS_check or ME_check)  and (qcpi>=6 or qper>=55) and (ugcpi>=ugcpi_l or ugper>=ugper_l) and (imper>=imper_l) and (hscpi>= hscpi_l or hsper>=hsper_l) and (gate_qualify or Exam_flag or gate_exemp)):
            check_flag3  = 1
            if(age<=37 or age_exemp):
                WrtFile(row,ind,sheet)
                # print('PAss1')
                row = row+1
            
                
            else:
                WrtFile(row_n,ind,sheet_n)
                # print('Fail1')
                row_n = row_n+1            
            
        elif(Btech_check  and (qcpi>=qcpi_btech or qper>=70)  and (ugcpi>=ugcpi_l or ugper>=ugper_l) and (imper>=imper_l) and (hscpi>= hscpi_l or hsper>=hsper_l) and (gate_valid or Exam_flag or gate_exemp)):
            check_flag2 = 1
            if(age<=33 or age_exemp):
                WrtFile(row,ind,sheet)
                # print('pass2')
                row = row+1
            else:
                check_flag = 1
                WrtFile(row_n,ind,sheet_n)
                # print('fail2')
                row_n = row_n+1  
        
        
        
        elif(MSC_check and (qcpi>=7 or qper>=65)  and (ugcpi>=ugcpi_l or ugper>=ugper_l) and (imper>=imper_l) and (hscpi>= hscpi_l or hsper>=hsper_l) and (gate_valid or Exam_flag or gate_exemp)):
            check_flag5 = 1
            if(age<=33 or age_exemp):
                WrtFile(row,ind,sheet)
                # print('pass3')
                row = row+1
            
                
            else:
                WrtFile(row_n,ind,sheet_n)
                # print('fail3')
                row_n = row_n+1              
    
        if(check_flag3 == 0 and (check_flag2 == 0 or check_flag == 1) and Btech_check and (df_in['214_n_degree_or_examination'][ind] != '' and type(df_in['214_n_degree_or_examination'][ind]) is not float)):
            check_flag4 = 1
            
            IIT_check = df_in['214_y_name_of_institution_or_university'][ind]
        
            if(df_in['214_e_percentage_of_marks_or_final_grade_point_average'][ind] != '' and pd.isna(df_in['214_e_percentage_of_marks_or_final_grade_point_average'][ind]) == False ):
                if(df_in['214_e_percentage_of_marks_or_final_grade_point_average'][ind]<=10):
                    qcpi = df_in['214_e_percentage_of_marks_or_final_grade_point_average'][ind]
                elif(df_in['214_e_percentage_of_marks_or_final_grade_point_average'][ind]>10):
                    qper = df_in['214_e_percentage_of_marks_or_final_grade_point_average'][ind]
            else:
                qcpi,qper = 0,0
        
            if(m.match(IIT_check)!= None or n.search(IIT_check)!=None or o.match(IIT_check)!= None):
       
                IIT_flag = 1          #Checking if IIT
            else:
                IIT_flag = 0
                
            if((qcpi>=6 or qper>=55) and (ugcpi>=ugcpi_l or ugper>=ugper_l) and (imper>=imper_l) and (hscpi>= hscpi_l or hsper>=hsper_l) and (gate_qualify or Exam_flag or gate_exemp)):
                if(age<=37 or age_exemp):
                    WrtFile(row,ind,sheet)
                    # print('PAss1_spec')
                    row = row+1
                    if(check_flag == 1):
                        sheet_n.delete_rows(row_n-1)
                        row_n-=1

                #print(Candidate, 'New' ,df_in['UserId'][ind] )
                else:
                    if(check_flag == 0):
                        WrtFile(row_n,ind,sheet_n)
                        # print('Fail1_special')
                        row_n = row_n+1 
            else:
                if(check_flag == 0):
                    WrtFile(row_n,ind,sheet_n)
                    # print('Fail2_special')
                    row_n = row_n+1 
        
        
        if(not(check_flag4) and not(check_flag3) and not(check_flag2) and not(check_flag5)):
            WrtFile(row_n,ind,sheet_n)
            # print('fail4')
            row_n = row_n+1   
    
    
    
    
    
    
    else:
        if((Mtech_check or MS_check or ME_check)  and (qcpi>=6.5 or qper>=60) and (ugcpi>=ugcpi_l or ugper>=ugper_l) and (imper>=imper_l) and (hscpi>= hscpi_l or hsper>=hsper_l) and (gate_qualify or Exam_flag or gate_exemp)):
            check_flag3 = 1
            if((category=='General' and (age<=32 or age_exemp)) or ((category=='OBC Non Creamy Layer'  or category== 'EWS'  or gender == 'Female'  or PD == 'Yes' ) and (age<=37 or age_exemp))):
                WrtFile(row,ind,sheet)
                # print('pass4')   
                row = row+1
            
                
            else:
                WrtFile(row_n,ind,sheet_n)
                # print('fail5')
                row_n = row_n+1            
            
        elif(Btech_check  and (qcpi>=qcpi_btech or qper>=75)  and (ugcpi>=ugcpi_l or ugper>=ugper_l) and (imper>=imper_l) and (hscpi>= hscpi_l or hsper>=hsper_l) and (gate_valid or Exam_flag or gate_exemp)):
            check_flag2 = 1
            if((category=='General' and (age<=28 or age_exemp)) or ((category=='OBC Non Creamy Layer'  or category== 'EWS'  or gender == 'Female'  or PD == 'Yes' ) and (age<=33 or age_exemp))):
                WrtFile(row,ind,sheet)
                # print('pass5')
                row = row+1
            else:
                check_flag = 1
                WrtFile(row_n,ind,sheet_n)
                # print('fail6')
                row_n = row_n+1  
            
        elif(MSC_check and (qcpi>=7.5 or qper>=70)  and (ugcpi>=ugcpi_l or ugper>=ugper_l) and (imper>=imper_l) and (hscpi>= hscpi_l or hsper>=hsper_l) and (gate_valid or Exam_flag or gate_exemp)):
            check_flag5 = 1
            if((category=='General' and (age<=28 or age_exemp)) or ((category=='OBC Non Creamy Layer'  or category== 'EWS'  or gender == 'Female'  or PD == 'Yes' ) and (age<=33 or age_exemp))):
                
                WrtFile(row,ind,sheet)
                # print('pass6')
                row = row+1  
    
            else:
                WrtFile(row_n,ind,sheet_n)
                # print('fail7')
                row_n = row_n+1  
        
        if(check_flag3 == 0 and (check_flag2 == 0 or check_flag == 1) and Btech_check and (df_in['214_n_degree_or_examination'][ind] != '' and type(df_in['214_n_degree_or_examination'][ind]) is not float)):
            check_flag4 = 1
            
            IIT_check = df_in['214_y_name_of_institution_or_university'][ind]
        
            if(df_in['214_e_percentage_of_marks_or_final_grade_point_average'][ind] != '' and pd.isna(df_in['214_e_percentage_of_marks_or_final_grade_point_average'][ind]) == False ):
                if(df_in['214_e_percentage_of_marks_or_final_grade_point_average'][ind]<=10):
                    qcpi = df_in['214_e_percentage_of_marks_or_final_grade_point_average'][ind]
                elif(df_in['214_e_percentage_of_marks_or_final_grade_point_average'][ind]>10):
                    qper = df_in['214_e_percentage_of_marks_or_final_grade_point_average'][ind]
            else:
                qcpi,qper = 0,0
        
            if(m.match(IIT_check)!= None or n.search(IIT_check)!=None or o.match(IIT_check)!= None):
       
                IIT_flag = 1          #Checking if IIT
            else:
                IIT_flag = 0
                
            if((qcpi>=6.5 or qper>=60) and (ugcpi>=ugcpi_l or ugper>=ugper_l) and (imper>=imper_l) and (hscpi>= hscpi_l or hsper>=hsper_l) and (gate_qualify or Exam_flag or gate_exemp)):
                if((category=='General' and (age<=32 or age_exemp)) or ((category=='OBC Non Creamy Layer'  or category== 'EWS'  or gender == 'Female'  or PD == 'Yes' ) and (age<=37 or age_exemp))):
                    WrtFile(row,ind,sheet)
                    # print('PAss1_spec_n')
                    row = row+1
                    if(check_flag == 1):
                        sheet_n.delete_rows(row_n-1)
                        row_n-=1

                #print(Candidate, 'New' ,df_in['UserId'][ind] )
                else:
                    if(check_flag == 0): 
                        WrtFile(row_n,ind,sheet_n)
                        # print('Fail1_special_n')
                        row_n = row_n+1 
            else:
                if(check_flag == 0):
                    WrtFile(row_n,ind,sheet_n)
                    # print('Fail2_special_n')
                    row_n = row_n+1 
        
        
        
        
        
        
        if(not(check_flag4) and not(check_flag3) and not(check_flag2) and not(check_flag5)):
            WrtFile(row_n,ind,sheet_n)
            # print('fail8')
            row_n = row_n+1
            
            
        
    if(df_in['UserId'][ind] == 38127 ):
       break
            
            
            
wb.save('Shortlisted.xlsx')                          #Saving to Excel         
wb_n.save('Not_shortlisted.xlsx')
rejected_df.to_excel('invalid_candidates.xlsx',index = False)    
print('Code Ran Successfully on' , input_file , 'The Files created are Shortlisted,Not_shortlisted and invalid candidates')  
for i in Invalid_degree:
    print('User ID',i[0],'Name',i[1],'No valid degree' , i[2])
       

        
        

        
        
        
    
    


      
    

   
    
    
    